# generate_sample_excel.py
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Mock data
report_data = [
    {
        "ID": "#101",
        "Mijoz": "Aliyev Vali",
        "Tel": "+998901234567",
        "Do'kon": "Chilonzor",
        "Buyurtma": "Taplyonniy: 2 dona, Yubileyniy: 1 dona",
        "Savdo (so'm)": 135000,
        "Tannarx (so'm)": 111000,
        "Foyda (so'm)": 24000,
        "Sana": "2026-05-24 10:30:15"
    },
    {
        "ID": "#102",
        "Mijoz": "Karimov Umid",
        "Tel": "+998935552233",
        "Do'kon": "Yunusobod",
        "Buyurtma": "Yulduz: 5 dona, Azbuka: 2 dona",
        "Savdo (so'm)": 380000,
        "Tannarx (so'm)": 332500,
        "Foyda (so'm)": 47500,
        "Sana": "2026-05-24 14:15:00"
    },
    {
        "ID": "#103",
        "Mijoz": "Sirojov Bobur",
        "Tel": "+998977778899",
        "Do'kon": "Qo'yliq",
        "Buyurtma": "Taplyonniy: 1 dona, Pop Corn: 3 dona",
        "Savdo (so'm)": 225000,
        "Tannarx (so'm)": 187500,
        "Foyda (so'm)": 37500,
        "Sana": "2026-05-24 18:45:22"
    }
]

df = pd.DataFrame(report_data)
filename = "sample_styled_report.xlsx"

with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Hisobot')
    workbook = writer.book
    worksheet = writer.sheets['Hisobot']
    
    # Gridlines
    worksheet.views.sheetView[0].showGridLines = True
    
    # Styles
    font_name = "Segoe UI"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    data_font = Font(name=font_name, size=10)
    total_font = Font(name=font_name, size=11, bold=True, color="000000")
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    border_double = Side(border_style="double", color="000000")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=Side(border_style="thin", color="000000"), bottom=border_double)
    
    fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Header formatting
    worksheet.row_dimensions[1].height = 26
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
        
    # Data formatting
    num_rows = len(df)
    for row_idx in range(2, num_rows + 2):
        worksheet.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            if is_even:
                cell.fill = fill_even
                
            # Alignments and number formats
            col_name = df.columns[col_idx - 1]
            if col_name in ["ID", "Tel", "Do'kon", "Sana"]:
                cell.alignment = align_center
            elif col_name in ["Mijoz", "Buyurtma"]:
                cell.alignment = align_left
            elif col_name in ["Savdo (so'm)", "Tannarx (so'm)", "Foyda (so'm)"]:
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
    # Total Row
    total_row_idx = num_rows + 2
    worksheet.row_dimensions[total_row_idx].height = 22
    
    # Label "JAMI:" in Column 1 (ID)
    cell_total_label = worksheet.cell(row=total_row_idx, column=1)
    cell_total_label.value = "JAMI:"
    cell_total_label.font = total_font
    cell_total_label.alignment = align_center
    cell_total_label.border = total_border
    
    # Empty values for intermediate columns with borders
    for col_idx in range(2, 6):
        cell = worksheet.cell(row=total_row_idx, column=col_idx)
        cell.value = ""
        cell.font = total_font
        cell.border = total_border
        
    # Formulas for numeric columns: 6 (Savdo), 7 (Tannarx), 8 (Foyda)
    for col_idx in [6, 7, 8]:
        col_letter = get_column_letter(col_idx)
        cell = worksheet.cell(row=total_row_idx, column=col_idx)
        cell.value = f"=SUM({col_letter}2:{col_letter}{num_rows+1})"
        cell.font = total_font
        cell.alignment = align_right
        cell.number_format = '#,##0'
        cell.border = total_border
        
    # Empty border for "Sana" column
    cell_sana = worksheet.cell(row=total_row_idx, column=9)
    cell_sana.value = ""
    cell_sana.border = total_border
    
    # Autofit columns
    for col in worksheet.columns:
        max_len = 0
        col_idx = col[0].column
        col_name = df.columns[col_idx - 1]
        
        for cell in col:
            val = str(cell.value or '')
            if cell.row == total_row_idx and '=' in val:
                # Don't measure raw formula string for autofit
                val = "1,000,000" 
            max_len = max(max_len, len(val))
            
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

print("Sample Excel created: sample_styled_report.xlsx")
