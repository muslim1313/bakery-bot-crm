import pandas as pd
import database as db
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

async def generate_excel_report(period='daily'):
    orders = await db.get_detailed_orders(period)
    
    if not orders:
        return None
    
    from config import PRODUCTS_PRICING
    import json
    
    report_data = []
    for o in orders:
        cart = json.loads(o['cart_json'])
        readable_cart = []
        for p_id, qty in cart.items():
            p_name = PRODUCTS_PRICING.get(p_id, {}).get('name', p_id)
            readable_cart.append(f"{p_name}: {qty} dona")
        
        cart_str = ", ".join(readable_cart)

        report_data.append({
            "ID": f"#{o['id']}",
            "Mijoz": o['name'],
            "Tel": o['phone'],
            "Do'kon": o.get('store', 'Noma\'lum'),
            "Buyurtma": cart_str,
            "Savdo (so'm)": o['total_revenue'],
            "Tannarx (so'm)": o['total_cost'],
            "Foyda (so'm)": o['profit'],
            "Sana": o['created_at']
        })
        
    df = pd.DataFrame(report_data)
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"report_{period}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hisobot')
        workbook = writer.book
        worksheet = writer.sheets['Hisobot']
        
        # Gridlines enabled
        worksheet.views.sheetView[0].showGridLines = True
        
        # Styling tokens
        font_name = "Segoe UI"
        header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        data_font = Font(name=font_name, size=10)
        total_font = Font(name=font_name, size=11, bold=True, color="000000")
        
        border_thin = Side(border_style="thin", color="D9D9D9")
        border_double = Side(border_style="double", color="000000")
        
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        total_border = Border(top=Side(border_style="thin", color="000000"), bottom=border_double)
        
        fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # Format Header row
        worksheet.row_dimensions[1].height = 26
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = header_border
            
        # Format Data rows
        num_rows = len(df)
        for row_idx in range(2, num_rows + 2):
            worksheet.row_dimensions[row_idx].height = 20
            is_even = (row_idx % 2 == 0)
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = cell_border
                if is_even:
                    cell.fill = fill_even
                    
                # Setup alignment & number format
                col_name = df.columns[col_idx - 1]
                if col_name in ["ID", "Tel", "Do'kon", "Sana"]:
                    cell.alignment = align_center
                elif col_name in ["Mijoz", "Buyurtma"]:
                    cell.alignment = align_left
                elif col_name in ["Savdo (so'm)", "Tannarx (so'm)", "Foyda (so'm)"]:
                    cell.alignment = align_right
                    cell.number_format = '#,##0'
                    
        # Add dynamic Total Row at the bottom
        total_row_idx = num_rows + 2
        worksheet.row_dimensions[total_row_idx].height = 22
        
        # "JAMI:" label
        cell_total_label = worksheet.cell(row=total_row_idx, column=1)
        cell_total_label.value = "JAMI:"
        cell_total_label.font = total_font
        cell_total_label.alignment = align_center
        cell_total_label.border = total_border
        
        # Border fillers
        for col_idx in range(2, 6):
            cell = worksheet.cell(row=total_row_idx, column=col_idx)
            cell.value = ""
            cell.font = total_font
            cell.border = total_border
            
        # Sum Formulas for Savdo (6), Tannarx (7), Foyda (8)
        for col_idx in [6, 7, 8]:
            col_letter = get_column_letter(col_idx)
            cell = worksheet.cell(row=total_row_idx, column=col_idx)
            cell.value = f"=SUM({col_letter}2:{col_letter}{num_rows+1})"
            cell.font = total_font
            cell.alignment = align_right
            cell.number_format = '#,##0'
            cell.border = total_border
            
        # Border filler for Sana (9)
        cell_sana = worksheet.cell(row=total_row_idx, column=9)
        cell_sana.value = ""
        cell_sana.border = total_border
        
        # Auto-fit Column Widths dynamically
        for col in worksheet.columns:
            max_len = 0
            col_idx = col[0].column
            
            for cell in col:
                val = str(cell.value or '')
                if cell.row == total_row_idx and '=' in val:
                    val = "1,000,000" # fallback length for formula
                max_len = max(max_len, len(val))
                
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    return filename
